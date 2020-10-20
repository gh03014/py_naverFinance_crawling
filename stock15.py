import requests 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import pyplot
from matplotlib import font_manager, rc
from matplotlib import font_manager as fm
import numpy as np
import re
from datetime import datetime
import os
import sys

class naver_finance:
    now = datetime.now()
    moon = str(now.month)
    today = str(now.day)
    today_date = (moon + "월" + today + "일")
    
    if(os.path.isfile('D:\\naver_finance.xlsx') == False):
        new_wb = Workbook()
        new_wb.save('D:\\naver_finance.xlsx')
    wb = load_workbook('D:\\naver_finance.xlsx')
    new_sheet = wb.create_sheet(today_date)
    new_sheet.append(['종목명', '현재가', '전일가', '전일대비', '시가', '고가', '저가', '마감시간'])

    def crawling(self):
        i = 0
        x = []
        y = []
        new_sheet = naver_finance.new_sheet
        self.get_price("DJI@DJI", x, y, new_sheet)
        self.get_price("LNS@FTSE100", x, y, new_sheet)
        self.get_price("NAS@IXIC", x, y, new_sheet)
        self.get_price("SPI@SPX", x, y, new_sheet)
        self.get_price("SHS@000001", x, y, new_sheet)
        self.get_price("HSI@HSI", x, y, new_sheet)
        self.get_price("PAS@CAC40", x, y, new_sheet)
        self.get_price("STX@SX5E", x, y, new_sheet)
        self.get_price("IDI@JKSE", x, y, new_sheet)
        self.get_price("NII@NI225", x, y, new_sheet)
        self.get_price("XTR@DAX30", x, y, new_sheet)
        self.get_price("BRI@BVSP", x, y, new_sheet)
        self.get_price("RUI@RTSI", x, y, new_sheet)
        self.get_price("MYI@KLSE", x, y, new_sheet)
        self.get_price("NAS@SOX", x, y, new_sheet)

        wb = naver_finance.wb
        wb.save('D:\\naver_finance.xlsx')
        
        fm.get_fontconfig_fonts()
        font_location = 'C:/Windows/Fonts/malgun.ttf'
        font_name = fm.FontProperties(fname = font_location).get_name()
        pyplot.rc('font', family=font_name)
        pyplot.rcParams["figure.figsize"] = (14, 7 )
        data_count = len(x)
        ypos = np.arange(data_count)
        rects = pyplot.barh(ypos, y, align = 'center', height = 0.7)
        pyplot.yticks(ypos, x)
        pyplot.xlabel('전일대비 변동량')
        pyplot.show()
        
    def graph_data_x(self, x):
        return x

    def graph_data_y(self, y):
        return y

    def get_bs_obj(self, company_code):
        url = "https://finance.naver.com/world/sise.nhn?symbol=" + company_code 
        result = requests.get(url) 
        bs_obj = BeautifulSoup(result.content, "html.parser") 
        return bs_obj

    def get_price(self, company_code, x, y, new_sheet):
        bs_obj = self.get_bs_obj(company_code)

        title = bs_obj.find("div", {"class": "group_h"})
        title2 = title.find("h2")
        title3 = title2.text.replace("\n", "")
        title4 = title3.replace(" ", "")
        title5 = ("종목:" + title4)

        price = bs_obj.find("p", {"class": "no_today"})
        price2 = price.find("em")
        price3 = price2.text.replace("\n", "")
        price4 = price3.replace(" ", "")
        price5 = "현재가:" + price4

        yester = bs_obj.find("table", {"class": "no_info"})
        yester2 = yester.find("em")
        yester3 = yester2.text.replace("\n", "")
        yester4 = yester3.replace(" ", "")
        yester5 = "전일가: " + yester4

        vary= bs_obj.find("p", {"class": "no_exday"})
        vary2 = vary.text.replace("\n", "")
        vary3 = vary2.strip("전일대비")
        vary4 = vary3.replace(" ", "")
        vary5 = "전일대비: " + vary4

        price_figure = price4.replace(",", "")
        price_figure2 = float(price_figure)
        yester_figure = yester4.replace(",", "")
        yester_figure2 = float(yester_figure)

        marketprice = bs_obj.find("table", {"class": "tb_status2 tb_status2_t2"})
        if(price_figure2 > yester_figure2):
            marketprice2 = marketprice.find("tr", {"class": "point_up"})
        else:
            marketprice2 = marketprice.find("tr", {"class": "point_dn"})
        marketprice3 = marketprice2.find("td", {"class": "tb_td4"})
        marketprice4 = marketprice3.text.replace("\n","")
        marketprice5 = marketprice4.replace(" ", "")
        marketprice6 = "시가: " + marketprice5

        highprice = marketprice2.find("td", {"class": "tb_td5"})
        highprice2 = highprice.text.replace("\n", "")
        highprice3 = highprice2.replace(" ", "")
        highprice4 = "고가: " + highprice3

        lowprice = marketprice2.find("td", {"class": "tb_td6"})
        lowprice2 = lowprice.text.replace("\n", "")
        lowprice3 = lowprice2.replace(" ", "")
        lowprice4 = "저가: " + lowprice3

        closetime = bs_obj.find("span", {"class": "date"})
        closetime2 = closetime.find("em")
        closetime3 = closetime2.text.replace("\n", "")
        closetime4 = closetime3.replace(" ", "-")
        closetime5 = "마감시간: " + closetime4

        if(price_figure2 > yester_figure2):
            vary_figure = price_figure2 - yester_figure2
        else:
            vary_figure = yester_figure2 - price_figure2
        vary_figure2 = round(vary_figure, 2)
        x.append(title4)
        y.append(vary_figure2)

        new_sheet.append([title4, price4, yester4, vary4, marketprice5, highprice3, lowprice3, closetime4])
        print(title5 + "  " + price5 + " " + yester5 + " " + vary5 + "\n"
               + marketprice6 + " " + highprice4 + " " + lowprice4 + " " + closetime5 + "\n")

vm = naver_finance()
vm.crawling()




